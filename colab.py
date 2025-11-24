import os
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins, PrintPageSetup
from datetime import datetime, timedelta
import yfinance as yf

# =======================================================
# إعدادات البروكسي والرؤوس (الجديدة)
# =======================================================

# رؤوس قياسية لمحاكاة متصفح حقيقي (لتجاوز الحظر)
STANDARD_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Language': 'en-US,en;q=0.9,ar;q=0.8',
    'Connection': 'keep-alive',
    'Referer': 'https://www.google.com/', 
}

# عنوان البروكسي للاختبار:
# قم بتعديل هذا السطر ببيانات البروكسي الحقيقية
# مثال: PROXY_ADDRESS = "http://user_abc:pass123@dc.oxylabs.io:8000"
PROXY_ADDRESS = "http://user-Davsa_RppK9-country-US:Aa+1020304050@dc.oxylabs.io:8000"
# ملاحظة: إذا لم يكن لديك بروكسي، اترك PROXY_ADDRESS كما هو للاتصال المباشر
# =======================================================

# تعريف المجموعات
SYMBOL_GROUPS = {
    '1': {
        'name': 'SPX-SPY-QQQ',
        'symbols': {
            'spx': '^GSPC',
            'spy': 'SPY',
            'qqq': 'QQQ',
        },
        'expiry_type': 'daily',
        'strike_count': 50
    },
    '2': {
        'name': 'Company',
        'symbols': {
           'AAPL': 'AAPL',
            'MSFT': 'MSFT',
            'NVDA': 'NVDA',
            'META': 'META',
            'AMZN': 'AMZN',
            'GOOGL': 'GOOGL',
            'AMD': 'AMD',
            'TSLA': 'TSLA',
            'ORCL': 'ORCL',
            'CRWD': 'CRWD',
            'COIN': 'COIN',
            'MSTR': 'MSTR',
            'AVGO': 'AVGO',
            'MU': 'MU',
            'CRM': 'CRM',
            'ARM': 'ARM',
            'CVNA': 'CVNA',
            'IBM': 'IBM',
            'BA': 'BA',
            'SMCI': 'SMCI',
            'RDDT': 'RDDT',
            'FSLR': 'FSLR',
            'RKLB': 'RKLB',
            'RBLX': 'RBLX',
            'PLTR': 'PLTR',
            'MCD': 'MCD',
            'NKE': 'NKE',
			 'LLY': 'LLY',
            'COST': 'COST',
            'APP': 'APP',
            'UBER': 'UBER',
            'XOM': 'XOM'
        },
        'expiry_type': 'weekly',
        'strike_count': 25
    }
}

# إعدادات التنسيق
BASE_FONT = Font(name='Calibri', size=9)
BOLD_FONT = Font(name='Calibri', size=9, bold=True)
HEADER_FONT = Font(name='Calibri', size=10, bold=True)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='e2e3e6'),
    right=Side(style='thin', color='e2e3e6'),
    top=Side(style='thin', color='e2e3e6'),
    bottom=Side(style='thin', color='e2e3e6')
)

# ألوان التعبئة
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL = PatternFill(start_color='F5CBCC', end_color='F5CBCC', fill_type='solid')
HEADER_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
PURPLE_FILL = PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

def get_hloc_data(symbol):
    """جلب بيانات HLOC من Yahoo Finance"""
    try:
        ticker = yf.Ticker(symbol)
        hist = ticker.history(period="1d")
        if hist.empty:
            return None, None, None, None, None
        return (
            round(hist['High'].iloc[-1], 2),
            round(hist['Low'].iloc[-1], 2),
            round(hist['Open'].iloc[-1], 2),
            round(hist['Close'].iloc[-1], 2),
            hist['Volume'].iloc[-1]
        )
    except Exception as e:
        print(f"فشل في جلب بيانات HLOC لـ {symbol}: {e}")
        return None, None, None, None, None

def download_options_data(symbol):
    """تنزيل بيانات الأوبشنز من OCC باستخدام بروكسي (إذا توفر)"""
    url = f"https://marketdata.theocc.com/series-search?symbolType=U&symbol={symbol}"
    
    # إعداد البروكسي ليتم تمريره إلى الطلب
    proxies = {
        "http": PROXY_ADDRESS,
        "https": PROXY_ADDRESS,
    }
    
    # تجاهل البروكسي إذا كان مجرد Placeholder
    if "YOUR_PROXY_ADDRESS" in PROXY_ADDRESS:
        proxies = None
    
    try:
        # استخدام الرؤوس والبروكسي في طلب requests.get
        response = requests.get(url, headers=STANDARD_HEADERS, proxies=proxies, timeout=60)
        
        if response.status_code == 403:
            print(f"❌ فشل: تم الحظر (403) حتى مع البروكسي.")
            return None
        
        return response.text if response.status_code == 200 else None
        
    except requests.exceptions.ProxyError:
        print(f"❌ فشل: خطأ اتصال بالبروكسي. تأكد من أن البروكسي يعمل.")
        return None
    except Exception as e:
        print(f"فشل في تنزيل بيانات لـ {symbol}: {e}")
        return None

def process_data(content, expiry_date, close_price, strike_count):
    """معالجة بيانات الأوبشنز واستخلاص السترايكات المحيطة بسعر الإغلاق"""
    try:
        lines = [line.strip() for line in content.split('\n') if line.strip()]
        start_idx = next((i for i, line in enumerate(lines) if line.startswith("ProductSymbol")), -1)

        if start_idx == -1:
            return pd.DataFrame(), []

        data = []
        seen_strikes = set()

        for line in lines[start_idx+1:]:
            parts = line.split()
            if len(parts) >= 11:
                try:
                    occ_symbol = parts[0]

                    if occ_symbol[0].isdigit():
                        continue

                    expiry = f"{parts[1]}-{parts[2].zfill(2)}-{parts[3].zfill(2)}"
                    if expiry != expiry_date:
                        continue

                    strike = float(f"{parts[4]}.{parts[5]}")
                    call_oi = int(parts[8])
                    put_oi = int(parts[9])

                    if strike not in seen_strikes:
                        data.append([strike, call_oi, put_oi])
                        seen_strikes.add(strike)
                except Exception as e:
                    print(f"خطأ في معالجة السطر: {line} - {e}")
                    continue

        if not data:
            return pd.DataFrame(), []

        df = pd.DataFrame(data, columns=["Strike", "Calls", "Puts"])
        df = df.sort_values('Strike')

        if close_price:
            upper = df[df['Strike'] >= close_price].nsmallest(strike_count, 'Strike')
            lower = df[df['Strike'] <= close_price].nlargest(strike_count, 'Strike')
            filtered_df = pd.concat([upper, lower]).drop_duplicates().sort_values('Strike')
            return filtered_df, filtered_df['Strike'].tolist()

        return df, df['Strike'].tolist()

    except Exception as e:
        print(f"خطأ في معالجة البيانات: {e}")
        return pd.DataFrame(), []

def get_expiry_date(group, today):
    """تحديد تاريخ الانتهاء بناءً على نوع المجموعة مع التعامل مع عطلات نهاية الأسبوع"""
    if group['expiry_type'] == 'daily':
        # تحقق من يوم الأسبوع (0=الإثنين، 1=الثلاثاء، ...، 5=السبت، 6=الأحد)
        weekday = today.weekday()
        if weekday == 5:  # السبت
            next_monday = today + timedelta(days=2)  # أضف يومين للوصول إلى الإثنين
            return next_monday.strftime("%Y-%m-%d")
        elif weekday == 6:  # الأحد
            next_monday = today + timedelta(days=1)  # أضف يومًا للوصول إلى الإثنين
            return next_monday.strftime("%Y-%m-%d")
        else:  # من الإثنين إلى الجمعة
            return today.strftime("%Y-%m-%d")  # استخدم التاريخ الحالي
    else:  # weekly
        next_friday = today + timedelta(days=(4 - today.weekday() + 7) % 7)
        return next_friday.strftime("%Y-%m-%d")  # الجمعة القادمة

def create_report(group_id):
    """إنشاء التقرير مع السترايكات المفلترة لمجموعة معينة"""
    group = SYMBOL_GROUPS[group_id]
    today = datetime.now().date()
    expiry_date = get_expiry_date(group, today)
    report_date = datetime.now().strftime("%Y%m%d")  # تاريخ اليوم بصيغة 20250714
    strike_count = group['strike_count']  # الحصول على عدد السترايكات من إعدادات المجموعة

    # إنشاء ملف Excel جديد لكل مجموعة
    wb = Workbook()
    if len(wb.sheetnames) == 1:
        wb.remove(wb.active)

    for symbol_name, symbol in group['symbols'].items():
        ws = wb.create_sheet(title=symbol_name[:31])

        print(f"Get== {symbol_name.upper()}...")

        high, low, open_, close, _ = get_hloc_data(symbol)

        content = download_options_data(symbol_name)
        if content is None:
            print(f"تخطي {symbol_name} بسبب نقص بيانات الأوبشنز")
            continue

        df, strikes = process_data(content, expiry_date, close, strike_count)
        if df.empty:
            print(f"لا توجد بيانات أوبشنز لـ {symbol_name}")
            continue

        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2)
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = True

        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        hloc_text = f"H: {high} | L: {low} | O: {open_} | C: {close}"
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=16)
        cell = ws.cell(row=1, column=1, value=hloc_text)
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = HEADER_FILL

        info_text = f"{symbol_name.upper()} | EXP: {expiry_date} | Report Date: {current_date}"
        ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=16)
        cell = ws.cell(row=2, column=1, value=info_text)
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = HEADER_FILL

        for col in range(1, 17):
            col_letter = get_column_letter(col)
            if col in [4, 8, 12, 16]:
                ws.column_dimensions[col_letter].width = 8
            else:
                ws.column_dimensions[col_letter].width = 8

        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 20
        for row in range(3, 200):
            ws.row_dimensions[row].height = 12

        start_col = 7
        data_start_row = 5

        titles = ["CALLS", "STRIKE", "PUTS"]
        for i, title in enumerate(titles):
            cell = ws.cell(row=4, column=start_col+i, value=title)
            cell.font = BOLD_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        for i, (_, row) in enumerate(df.iterrows()):
            # كتابة البيانات في الخلايا مع الحفاظ على القيم الرقمية
            values = [row['Calls'], row['Strike'], row['Puts']]
            
            for j, value in enumerate(values):
                cell = ws.cell(row=data_start_row+i, column=start_col+j, value=value)
                cell.font = BASE_FONT
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER
                
                # تطبيق تنسيق الأرقام
                if j == 0 or j == 2:  # Calls و Puts
                    cell.number_format = '#,##0'  # تنسيق بفاصلة الآلاف بدون منازل عشرية
                elif j == 1:  # Strike
                    cell.number_format = '0'   # تنسيق بدون منازل عشرية وبدون فاصلة الآلاف

        highlight_close_price(ws, start_col, data_start_row, close, strikes)
        highlight_top5_oi(ws, start_col, data_start_row, close, strikes)

    filename = f"{group['name'].replace('-', '_')}_{report_date}.xlsx"
    wb.save(filename)
    print(f"\nتم إنشاء التقرير بنجاح: {filename}")
    print(f"تم حفظ الملف في: {os.path.abspath(filename)}")

def highlight_close_price(ws, start_col, start_row, close_price, strikes):
    """إبراز خلايا السترايك القريبة من سعر الإغلاق باللون الأصفر"""
    if close_price is None or not strikes:
        return

    sorted_strikes = sorted(strikes, key=lambda x: abs(x - close_price))
    closest_strikes = sorted_strikes[:2]

    for row in range(start_row, start_row + len(strikes)):
        strike = ws.cell(row=row, column=start_col+1).value
        if strike is None:
            continue

        try:
            strike = float(strike)

            if strike in closest_strikes:
                cell = ws.cell(row=row, column=start_col+1)
                cell.fill = YELLOW_FILL
                cell.font = BOLD_FONT

        except (ValueError, TypeError):
            continue

def highlight_top5_oi(ws, start_col, start_row, close_price, strikes):
    """إبراز أعلى 5 قيم OI للـ Calls وPuts خارج النطاق"""
    if not close_price or not strikes:
        return

    calls_otm = []
    puts_otm = []

    for row in range(start_row, start_row + len(strikes)):
        strike = ws.cell(row=row, column=start_col + 1).value
        if strike is None:
            continue

        try:
            strike = float(strike)
            # الحصول على القيمة الرقمية مباشرة (لأنها محفوظة كأرقام)
            call_oi = ws.cell(row=row, column=start_col).value or 0
            put_oi = ws.cell(row=row, column=start_col + 2).value or 0

            if strike > close_price:
                calls_otm.append({'row': row, 'value': call_oi})
            elif strike < close_price:
                puts_otm.append({'row': row, 'value': put_oi})

        except (ValueError, TypeError):
            continue

    calls_otm_sorted = sorted(calls_otm, key=lambda x: -x['value'])[:5]
    puts_otm_sorted = sorted(puts_otm, key=lambda x: -x['value'])[:5]

    light_green = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    for item in calls_otm_sorted:
        ws.cell(row=item['row'], column=start_col).fill = light_green

    light_pink = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    for item in puts_otm_sorted:
        ws.cell(row=item['row'], column=start_col + 2).fill = light_pink

if __name__ == "__main__":
    # تشغيل التقرير لكل مجموعة
    for group_id in ['1', '2']:
        create_report(group_id)
